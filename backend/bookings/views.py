from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Sum
from django.utils import timezone
from datetime import date
from .models import Booking, User, Room, RoomIssue
from .serializers import (
    BookingSerializer, BookingListSerializer, UserSerializer, 
    RoomSerializer, RoomAvailabilitySerializer, RoomIssueSerializer
)
from .permissions import IsManagerOrReadOnly, IsManager


class BookingViewSet(viewsets.ModelViewSet):
    serializer_class = BookingSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        queryset = Booking.objects.filter(is_original=True)
        
        # Filter by check-in date if provided
        check_in_date = self.request.query_params.get('check_in_date', None)
        if check_in_date:
            try:
                queryset = queryset.filter(check_in_date=check_in_date)
            except ValueError:
                pass
        
        # Receptionist only sees authorized bookings, manager sees all
        if user.is_receptionist():
            queryset = queryset.filter(is_authorized=True)
        
        # Exclude soft-deleted bookings by default (unless manager explicitly requests them)
        include_deleted = self.request.query_params.get('include_deleted', 'false').lower() == 'true'
        if not include_deleted or user.is_receptionist():
            queryset = queryset.filter(deleted_at__isnull=True)
        
        return queryset.order_by('-created_at')
    
    def get_serializer_class(self):
        if self.action == 'list':
            return BookingListSerializer
        return BookingSerializer
    
    def get_permissions(self):
        if self.action == 'destroy':
            # Only manager can delete
            return [IsAuthenticated(), IsManager()]
        elif self.action in ['update', 'partial_update']:
            # Both can edit, but manager edits supersede
            return [IsAuthenticated()]
        return [IsAuthenticated()]
    
    def perform_create(self, serializer):
        serializer.save(booked_by=self.request.user)
    
    def perform_update(self, serializer):
        # The update logic is handled in the serializer
        serializer.save()
    
    def perform_destroy(self, instance):
        # Only manager can delete (soft delete)
        if not self.request.user.is_manager():
            return Response(
                {"detail": "You do not have permission to delete bookings."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Soft delete instead of hard delete
        from django.utils import timezone
        instance.deleted_at = timezone.now()
        instance.deleted_by = self.request.user
        instance.save()
    
    @action(detail=True, methods=['get'])
    def versions(self, request, pk=None):
        """Get all versions of a booking (manager only)"""
        if not request.user.is_manager():
            return Response(
                {"detail": "Only managers can view booking versions."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        booking = self.get_object()
        versions = booking.versions.all()
        
        from .serializers import BookingVersionSerializer
        serializer = BookingVersionSerializer(versions, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def export_versions_excel(self, request, pk=None):
        """Export booking versions to Excel file (manager only)"""
        if not request.user.is_manager():
            return Response(
                {"detail": "Only managers can export booking versions."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from django.utils import timezone
        import json
        
        booking = self.get_object()
        versions = booking.versions.all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Booking Versions"
        
        # Headers
        headers = [
            'Version ID', 'Booking ID', 'Guest Name', 'Room Number',
            'Edited By', 'Edited At', 'Is Manager Edit', 'Version Data (JSON)'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, version in enumerate(versions, 2):
            version_data = version.version_data
            ws.cell(row=row_num, column=1, value=version.id)
            ws.cell(row=row_num, column=2, value=booking.id)
            ws.cell(row=row_num, column=3, value=version_data.get('name', ''))
            ws.cell(row=row_num, column=4, value=version_data.get('room_no', ''))
            ws.cell(row=row_num, column=5, value=version.edited_by.username if version.edited_by else '')
            ws.cell(row=row_num, column=6, value=version.edited_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=7, value='Yes' if version.is_manager_edit else 'No')
            ws.cell(row=row_num, column=8, value=json.dumps(version_data, indent=2))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"booking_versions_{booking.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def daily_totals(self, request):
        """Get daily totals for bookings"""
        target_date = request.query_params.get('date', None)
        
        if target_date:
            try:
                target_date = date.fromisoformat(target_date)
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            target_date = timezone.now().date()
        
        bookings = Booking.objects.filter(
            is_original=True,
            check_in_date=target_date,
            deleted_at__isnull=True  # Exclude soft-deleted bookings
        )
        
        total_amount = bookings.aggregate(total=Sum('amount_ghs'))['total'] or 0
        total_bookings = bookings.count()
        
        return Response({
            'date': target_date,
            'total_bookings': total_bookings,
            'total_amount_ghs': float(total_amount)
        })
    
    @action(detail=True, methods=['post'])
    def authorize(self, request, pk=None):
        """Authorize a booking (manager only)"""
        if not request.user.is_manager():
            return Response(
                {"detail": "Only managers can authorize bookings."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        booking = self.get_object()
        authorized_by = request.data.get('authorized_by', request.user.get_full_name() or request.user.username)
        
        booking.is_authorized = True
        booking.authorized_by = authorized_by
        booking.save()
        
        serializer = self.get_serializer(booking)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore a soft-deleted booking (manager only, within 30 days)"""
        if not request.user.is_manager():
            return Response(
                {"detail": "Only managers can restore bookings."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        booking = self.get_object()
        
        if not booking.is_deleted:
            return Response(
                {"detail": "This booking is not deleted."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not booking.can_restore:
            return Response(
                {"detail": "This booking cannot be restored. The 30-day restoration period has expired."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        booking.deleted_at = None
        booking.deleted_by = None
        booking.save()
        
        serializer = self.get_serializer(booking)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Export bookings to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Get bookings based on user role
        queryset = self.get_queryset()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Headers
        headers = [
            'ID', 'Guest Name', 'ID/Telephone', 'Address/Location', 'Age',
            'Room Number', 'Room Type', 'Check-in Date', 'Check-in Time',
            'Check-out Date', 'Check-out Time', 'Payment Method',
            'Amount (GHS)', 'Cash Amount', 'MoMo Amount', 'MoMo Network',
            'MoMo Number', 'Booked By', 'Created At', 'Last Edited By',
            'Updated At', 'Authorized By', 'Is Authorized', 'Version Number'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, booking in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=booking.id)
            ws.cell(row=row_num, column=2, value=booking.name)
            ws.cell(row=row_num, column=3, value=booking.id_or_telephone)
            ws.cell(row=row_num, column=4, value=booking.address_location)
            ws.cell(row=row_num, column=5, value=booking.age)
            ws.cell(row=row_num, column=6, value=booking.room_no)
            ws.cell(row=row_num, column=7, value=booking.room.get_room_type_display() if booking.room else 'N/A')
            ws.cell(row=row_num, column=8, value=str(booking.check_in_date))
            ws.cell(row=row_num, column=9, value=str(booking.check_in_time))
            ws.cell(row=row_num, column=10, value=str(booking.check_out_date) if booking.check_out_date else '')
            ws.cell(row=row_num, column=11, value=str(booking.check_out_time) if booking.check_out_time else '')
            ws.cell(row=row_num, column=12, value=booking.get_payment_method_display())
            ws.cell(row=row_num, column=13, value=float(booking.amount_ghs))
            ws.cell(row=row_num, column=14, value=float(booking.cash_amount))
            ws.cell(row=row_num, column=15, value=float(booking.momo_amount))
            ws.cell(row=row_num, column=16, value=booking.momo_network or '')
            ws.cell(row=row_num, column=17, value=booking.momo_number)
            ws.cell(row=row_num, column=18, value=booking.booked_by.username if booking.booked_by else '')
            ws.cell(row=row_num, column=19, value=booking.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=20, value=booking.last_edited_by.username if booking.last_edited_by else '')
            ws.cell(row=row_num, column=21, value=booking.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=22, value=booking.authorized_by)
            ws.cell(row=row_num, column=23, value='Yes' if booking.is_authorized else 'No')
            ws.cell(row=row_num, column=24, value=booking.version_number)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"bookings_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def available(self, request):
        """Get available rooms for given dates"""
        check_in_date = request.query_params.get('check_in_date', None)
        check_out_date = request.query_params.get('check_out_date', None)
        room_type = request.query_params.get('room_type', None)
        
        if not check_in_date:
            return Response(
                {"error": "check_in_date parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            check_in = date.fromisoformat(check_in_date)
            check_out = date.fromisoformat(check_out_date) if check_out_date else None
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        rooms = Room.objects.filter(is_available=True)
        
        if room_type:
            rooms = rooms.filter(room_type=room_type)
        
        available_rooms = []
        for room in rooms:
            is_available = room.check_availability(check_in, check_out)
            if is_available:
                available_rooms.append({
                    'room_id': room.id,
                    'room_number': room.room_number,
                    'room_type': room.room_type,
                    'room_type_display': room.get_room_type_display(),
                    'is_available': True,
                    'price_per_night': float(room.price_per_night),
                    'description': room.description
                })
        
        return Response(available_rooms)
    
    @action(detail=False, methods=['get'])
    def status(self, request):
        """Get all rooms with their booking status for a given date"""
        target_date = request.query_params.get('date', None)
        
        if target_date:
            try:
                target_date = date.fromisoformat(target_date)
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            target_date = timezone.now().date()
        
        rooms = Room.objects.all()
        room_status = []
        
        for room in rooms:
            is_booked = not room.check_availability(target_date)
            current_booking = None
            if is_booked:
                booking = Booking.objects.filter(
                    room_no=room.room_number,
                    is_original=True,
                    check_in_date__lte=target_date
                ).exclude(
                    check_out_date__lt=target_date
                ).first()
                if booking:
                    current_booking = {
                        'guest_name': booking.name,
                        'check_in': str(booking.check_in_date),
                        'check_out': str(booking.check_out_date) if booking.check_out_date else None
                    }
            
            room_status.append({
                'room_id': room.id,
                'room_number': room.room_number,
                'room_type': room.room_type,
                'room_type_display': room.get_room_type_display(),
                'is_available': room.is_available,
                'is_booked': is_booked,
                'current_booking': current_booking,
                'price_per_night': float(room.price_per_night)
            })
        
        return Response(room_status)


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user information"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)


class RoomIssueViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing room issues, faults, and missing inventory.
    Both managers and receptionists have full access to:
    - View all issues
    - Create new issues
    - Update issues
    - Mark issues as fixed
    - View summaries and statistics
    """
    queryset = RoomIssue.objects.all()
    serializer_class = RoomIssueSerializer
    permission_classes = [IsAuthenticated]  # Both managers and receptionists have equal access
    
    def get_queryset(self):
        # Both managers and receptionists can see all issues
        queryset = RoomIssue.objects.all()
        
        # Filter by room if provided
        room_id = self.request.query_params.get('room', None)
        if room_id:
            queryset = queryset.filter(room_id=room_id)
        
        # Filter by status if provided
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by issue_type if provided
        issue_type = self.request.query_params.get('issue_type', None)
        if issue_type:
            queryset = queryset.filter(issue_type=issue_type)
        
        # Filter unresolved issues
        unresolved = self.request.query_params.get('unresolved', None)
        if unresolved and unresolved.lower() == 'true':
            queryset = queryset.exclude(status__in=['fixed', 'resolved'])
        
        return queryset.order_by('-reported_at')
    
    @action(detail=True, methods=['post'])
    def mark_fixed(self, request, pk=None):
        """Mark an issue as fixed"""
        issue = self.get_object()
        resolution_notes = request.data.get('resolution_notes', '')
        
        issue.mark_as_fixed(request.user, notes=resolution_notes)
        
        serializer = self.get_serializer(issue)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_room(self, request):
        """Get all issues for a specific room"""
        room_id = request.query_params.get('room_id', None)
        if not room_id:
            return Response(
                {"error": "room_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        issues = RoomIssue.objects.filter(room_id=room_id).order_by('-reported_at')
        serializer = self.get_serializer(issues, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get summary of room issues"""
        total_issues = RoomIssue.objects.count()
        unresolved = RoomIssue.objects.exclude(status__in=['fixed', 'resolved']).count()
        
        by_type = {}
        for issue_type, display in RoomIssue.ISSUE_TYPE_CHOICES:
            by_type[issue_type] = {
                'display': display,
                'count': RoomIssue.objects.filter(issue_type=issue_type).count(),
                'unresolved': RoomIssue.objects.filter(
                    issue_type=issue_type
                ).exclude(status__in=['fixed', 'resolved']).count()
            }
        
        by_status = {}
        for status_val, display in RoomIssue.STATUS_CHOICES:
            by_status[status_val] = {
                'display': display,
                'count': RoomIssue.objects.filter(status=status_val).count()
            }
        
        return Response({
            'total_issues': total_issues,
            'unresolved': unresolved,
            'resolved': total_issues - unresolved,
            'by_type': by_type,
            'by_status': by_status
        })
    
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Export room issues to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Get issues based on filters
        queryset = self.get_queryset()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Room Issues"
        
        # Headers
        headers = [
            'ID', 'Room Number', 'Room Type', 'Issue Type', 'Title',
            'Description', 'Status', 'Priority', 'Reported By',
            'Reported At', 'Fixed By', 'Fixed At', 'Resolution Notes',
            'Created At', 'Updated At'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, issue in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=issue.id)
            ws.cell(row=row_num, column=2, value=issue.room.room_number)
            ws.cell(row=row_num, column=3, value=issue.room.get_room_type_display())
            ws.cell(row=row_num, column=4, value=issue.get_issue_type_display())
            ws.cell(row=row_num, column=5, value=issue.title)
            ws.cell(row=row_num, column=6, value=issue.description)
            ws.cell(row=row_num, column=7, value=issue.get_status_display())
            ws.cell(row=row_num, column=8, value=issue.get_priority_display())
            ws.cell(row=row_num, column=9, value=issue.reported_by.username if issue.reported_by else '')
            ws.cell(row=row_num, column=10, value=issue.reported_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=11, value=issue.fixed_by.username if issue.fixed_by else '')
            ws.cell(row=row_num, column=12, value=issue.fixed_at.strftime('%Y-%m-%d %H:%M:%S') if issue.fixed_at else '')
            ws.cell(row=row_num, column=13, value=issue.resolution_notes)
            ws.cell(row=row_num, column=14, value=issue.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=15, value=issue.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"room_issues_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

