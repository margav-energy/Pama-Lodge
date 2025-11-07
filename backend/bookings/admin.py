from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.forms import UserCreationForm
from django import forms
from django.utils import timezone
from django.utils.html import format_html
from django.db.models import Count, Sum, Q
from django.utils import timezone as tz
from django.http import HttpResponse
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import User, Booking, BookingVersion, Room, RoomIssue

# Customize admin site header and title
admin.site.site_header = 'Pama Lodge Administration'
admin.site.site_title = 'Pama Lodge Administration'
admin.site.index_title = 'Welcome to Pama Lodge Administration'


# Override admin index to add statistics
original_index = admin.site.index

def custom_index(request, extra_context=None):
    """Custom admin index with statistics"""
    extra_context = extra_context or {}
    
    # Get statistics
    today = tz.now().date()
    week_ago = today - timedelta(days=7)
    
    # Booking statistics
    total_bookings = Booking.objects.filter(is_original=True, deleted_at__isnull=True).count()
    authorized_bookings = Booking.objects.filter(is_original=True, status='authorized', deleted_at__isnull=True).count()
    pending_bookings = Booking.objects.filter(is_original=True, status='pending', deleted_at__isnull=True).count()
    rejected_bookings = Booking.objects.filter(is_original=True, status='rejected', deleted_at__isnull=True).count()
    deleted_bookings = Booking.objects.filter(is_original=True, deleted_at__isnull=False).count()
    today_bookings = Booking.objects.filter(is_original=True, check_in_date=today, deleted_at__isnull=True).count()
    week_bookings = Booking.objects.filter(is_original=True, check_in_date__gte=week_ago, deleted_at__isnull=True).count()
    
    # Revenue statistics
    total_revenue = Booking.objects.filter(
        is_original=True, 
        status='authorized', 
        deleted_at__isnull=True
    ).aggregate(total=Sum('amount_ghs'))['total'] or 0
    
    today_revenue = Booking.objects.filter(
        is_original=True,
        check_in_date=today,
        status='authorized',
        deleted_at__isnull=True
    ).aggregate(total=Sum('amount_ghs'))['total'] or 0
    
    week_revenue = Booking.objects.filter(
        is_original=True,
        check_in_date__gte=week_ago,
        status='authorized',
        deleted_at__isnull=True
    ).aggregate(total=Sum('amount_ghs'))['total'] or 0
    
    # Room statistics
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(is_available=True).count()
    booked_rooms = total_rooms - available_rooms
    
    # User statistics
    total_users = User.objects.count()
    managers = User.objects.filter(role='manager').count()
    receptionists = User.objects.filter(role='receptionist').count()
    
    # Room Issues statistics
    total_issues = RoomIssue.objects.count()
    reported_issues = RoomIssue.objects.filter(status='reported').count()
    in_progress_issues = RoomIssue.objects.filter(status='in_progress').count()
    fixed_issues = RoomIssue.objects.filter(status='fixed').count()
    
    # Booking Versions
    total_versions = BookingVersion.objects.count()
    
    extra_context.update({
        'stats': {
            'bookings': {
                'total': total_bookings,
                'authorized': authorized_bookings,
                'pending': pending_bookings,
                'rejected': rejected_bookings,
                'deleted': deleted_bookings,
                'today': today_bookings,
                'this_week': week_bookings,
            },
            'revenue': {
                'total': float(total_revenue),
                'today': float(today_revenue),
                'this_week': float(week_revenue),
            },
            'rooms': {
                'total': total_rooms,
                'available': available_rooms,
                'booked': booked_rooms,
            },
            'users': {
                'total': total_users,
                'managers': managers,
                'receptionists': receptionists,
            },
            'issues': {
                'total': total_issues,
                'reported': reported_issues,
                'in_progress': in_progress_issues,
                'fixed': fixed_issues,
            },
            'versions': {
                'total': total_versions,
            }
        }
    })
    
    return original_index(request, extra_context)

admin.site.index = custom_index


class CustomUserCreationForm(UserCreationForm):
    """Custom form that doesn't require username during creation"""
    class Meta:
        model = User
        fields = ('first_name', 'last_name', 'email', 'role')
        field_classes = {}
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Make username optional and hide it
        if 'username' in self.fields:
            self.fields['username'].required = False
            self.fields['username'].widget = forms.HiddenInput()
    
    def clean_username(self):
        # Return empty string since we'll generate it
        return ''


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    add_form = CustomUserCreationForm
    list_display = ['username', 'first_name', 'last_name', 'email', 'role', 'is_staff', 'date_joined']
    list_filter = ['role', 'is_staff', 'is_superuser']
    actions = ['export_excel']
    change_list_template = 'admin/bookings/user/change_list.html'
    
    # Fieldsets for editing existing users
    fieldsets = BaseUserAdmin.fieldsets + (
        ('Role', {'fields': ('role',)}),
    )
    
    # Fieldsets for creating new users - show first_name, last_name, hide username
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('first_name', 'last_name', 'email', 'password1', 'password2', 'role'),
        }),
    )
    
    def save_model(self, request, obj, form, change):
        # If creating a new user (not editing), auto-generate username
        if not change and not obj.username:
            # Generate username from first_name and last_name
            first_name = obj.first_name.strip().lower() if obj.first_name else ''
            last_name = obj.last_name.strip().lower() if obj.last_name else ''
            
            if first_name and last_name:
                # Replace spaces and special characters with underscores
                first_name = first_name.replace(' ', '_')
                last_name = last_name.replace(' ', '_')
                
                base_username = f"{first_name}_{last_name}"
                username = base_username
                counter = 1
                
                # Check if username already exists, if so add a number
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1
                
                obj.username = username
        
        super().save_model(request, obj, form, change)
    
    def export_excel(self, request, queryset):
        """Export users to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers
        headers = [
            'ID', 'Username', 'First Name', 'Last Name', 'Email',
            'Role', 'Is Staff', 'Is Superuser', 'Is Active', 'Date Joined', 'Last Login'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, user in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.username)
            ws.cell(row=row_num, column=3, value=user.first_name or '')
            ws.cell(row=row_num, column=4, value=user.last_name or '')
            ws.cell(row=row_num, column=5, value=user.email or '')
            ws.cell(row=row_num, column=6, value=user.get_role_display())
            ws.cell(row=row_num, column=7, value='Yes' if user.is_staff else 'No')
            ws.cell(row=row_num, column=8, value='Yes' if user.is_superuser else 'No')
            ws.cell(row=row_num, column=9, value='Yes' if user.is_active else 'No')
            ws.cell(row=row_num, column=10, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '')
            ws.cell(row=row_num, column=11, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"users_export_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected users to Excel"


@admin.register(Room)
class RoomAdmin(admin.ModelAdmin):
    list_display = ['room_number', 'room_type', 'price_per_night', 'is_available', 'created_at']
    list_filter = ['room_type', 'is_available']
    search_fields = ['room_number']
    list_editable = ['is_available']
    actions = ['export_excel']
    change_list_template = 'admin/bookings/room/change_list.html'
    
    def export_excel(self, request, queryset):
        """Export rooms to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rooms"
        
        # Headers
        headers = [
            'ID', 'Room Number', 'Room Type', 'Price Per Night (GHS)',
            'Is Available', 'Created At', 'Updated At'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, room in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=room.id)
            ws.cell(row=row_num, column=2, value=room.room_number)
            ws.cell(row=row_num, column=3, value=room.get_room_type_display())
            ws.cell(row=row_num, column=4, value=float(room.price_per_night))
            ws.cell(row=row_num, column=5, value='Yes' if room.is_available else 'No')
            ws.cell(row=row_num, column=6, value=room.created_at.strftime('%Y-%m-%d %H:%M:%S') if room.created_at else '')
            ws.cell(row=row_num, column=7, value=room.updated_at.strftime('%Y-%m-%d %H:%M:%S') if room.updated_at else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"rooms_export_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected rooms to Excel"


@admin.register(Booking)
class BookingAdmin(admin.ModelAdmin):
    list_display = ['name', 'room_no', 'check_in_date', 'amount_ghs', 'booked_by', 'status', 'is_deleted_display', 'created_at']
    list_filter = ['check_in_date', 'payment_method', 'status', 'deleted_at']
    search_fields = ['name', 'room_no', 'id_or_telephone']
    actions = ['soft_delete_selected', 'restore_selected', 'export_excel']
    change_list_template = 'admin/bookings/booking/change_list.html'
    
    def get_actions(self, request):
        """Remove default delete action and keep only soft delete"""
        actions = super().get_actions(request)
        if 'delete_selected' in actions:
            del actions['delete_selected']
        return actions
    
    def is_deleted_display(self, obj):
        """Display deleted status"""
        if obj.is_deleted:
            if obj.deleted_at:
                deleted_date = obj.deleted_at.strftime('%b %d, %Y')
                return f"Deleted ({deleted_date})"
            return "Deleted"
        return "Active"
    is_deleted_display.short_description = "Status"
    is_deleted_display.admin_order_field = 'deleted_at'
    
    def get_queryset(self, request):
        """Include deleted bookings in admin"""
        qs = super().get_queryset(request)
        return qs.filter(is_original=True)
    
    def soft_delete_selected(self, request, queryset):
        """Soft delete selected bookings"""
        # Only soft delete bookings that aren't already deleted
        to_delete = queryset.filter(deleted_at__isnull=True)
        count = to_delete.count()
        
        for booking in to_delete:
            booking.deleted_at = timezone.now()
            booking.deleted_by = request.user
            booking.save()
        
        if count > 0:
            self.message_user(request, f'{count} booking(s) soft deleted successfully.')
        else:
            self.message_user(request, 'No bookings to delete (all selected are already deleted).', level='warning')
    soft_delete_selected.short_description = "Soft delete selected bookings"
    
    def restore_selected(self, request, queryset):
        """Restore soft-deleted bookings"""
        # Only restore bookings that are deleted and within 30 days
        to_restore = queryset.filter(deleted_at__isnull=False)
        restored_count = 0
        expired_count = 0
        
        for booking in to_restore:
            if booking.can_restore:
                booking.deleted_at = None
                booking.deleted_by = None
                booking.save()
                restored_count += 1
            else:
                expired_count += 1
        
        messages = []
        if restored_count > 0:
            messages.append(f'{restored_count} booking(s) restored successfully.')
        if expired_count > 0:
            messages.append(f'{expired_count} booking(s) could not be restored (30-day period expired).')
        
        if messages:
            self.message_user(request, ' '.join(messages))
        else:
            self.message_user(request, 'No deleted bookings selected to restore.', level='warning')
    restore_selected.short_description = "Restore selected deleted bookings"
    
    def delete_model(self, request, obj):
        """Override delete to use soft delete"""
        obj.deleted_at = timezone.now()
        obj.deleted_by = request.user
        obj.save()
        self.message_user(request, f'Booking "{obj.name}" has been soft deleted.')
    
    def delete_queryset(self, request, queryset):
        """Override bulk delete to use soft delete"""
        count = queryset.filter(deleted_at__isnull=True).count()
        for booking in queryset.filter(deleted_at__isnull=True):
            booking.deleted_at = timezone.now()
            booking.deleted_by = request.user
            booking.save()
        
        if count > 0:
            self.message_user(request, f'{count} booking(s) soft deleted successfully.')
    
    def export_excel(self, request, queryset):
        """Export bookings to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Headers
        headers = [
            'ID', 'Guest Name', 'Phone Number', 'Address/Location', 'Age',
            'Room Number', 'Room Type', 'Check-in Date', 'Check-in Time',
            'Check-out Date', 'Check-out Time', 'Payment Method',
            'Amount (GHS)', 'Cash Amount', 'MoMo Amount', 'MoMo Network',
            'MoMo Number', 'Booked By', 'Created At', 'Last Edited By',
            'Updated At', 'Authorized/Rejected By', 'Status', 'Rejection Reason', 'Version Number'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, booking in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=booking.id)
            ws.cell(row=row_num, column=2, value=booking.name)
            ws.cell(row=row_num, column=3, value=booking.id_or_telephone)
            ws.cell(row=row_num, column=4, value=booking.address_location)
            ws.cell(row=row_num, column=5, value=booking.age)
            ws.cell(row=row_num, column=6, value=booking.room_no)
            ws.cell(row=row_num, column=7, value=booking.room.get_room_type_display() if booking.room else 'N/A')
            ws.cell(row=row_num, column=8, value=str(booking.check_in_date))
            ws.cell(row=row_num, column=9, value=str(booking.check_in_time))
            ws.cell(row=row_num, column=10, value=str(booking.check_out_date) if booking.check_out_date else '')
            ws.cell(row=row_num, column=11, value=str(booking.check_out_time) if booking.check_out_time else '')
            ws.cell(row=row_num, column=12, value=booking.get_payment_method_display())
            ws.cell(row=row_num, column=13, value=float(booking.amount_ghs))
            ws.cell(row=row_num, column=14, value=float(booking.cash_amount))
            ws.cell(row=row_num, column=15, value=float(booking.momo_amount))
            ws.cell(row=row_num, column=16, value=booking.momo_network or '')
            ws.cell(row=row_num, column=17, value=booking.momo_number)
            ws.cell(row=row_num, column=18, value=booking.booked_by.username if booking.booked_by else '')
            ws.cell(row=row_num, column=19, value=booking.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=20, value=booking.last_edited_by.username if booking.last_edited_by else '')
            ws.cell(row=row_num, column=21, value=booking.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=22, value=booking.authorized_by or booking.rejected_by or '')
            ws.cell(row=row_num, column=23, value=booking.get_status_display())
            ws.cell(row=row_num, column=24, value=booking.rejection_reason or '')
            ws.cell(row=row_num, column=25, value=booking.version_number)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"bookings_export_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected bookings to Excel"


@admin.register(BookingVersion)
class BookingVersionAdmin(admin.ModelAdmin):
    list_display = ['booking', 'edited_by', 'edited_at', 'is_manager_edit']
    list_filter = ['is_manager_edit', 'edited_at']
    readonly_fields = ['edited_at']
    actions = ['export_excel']
    change_list_template = 'admin/bookings/bookingversion/change_list.html'
    
    def export_excel(self, request, queryset):
        """Export booking versions to Excel file"""
        import json
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Booking Versions"
        
        # Headers
        headers = [
            'Version ID', 'Booking ID', 'Guest Name', 'Room Number',
            'Edited By', 'Edited At', 'Is Manager Edit', 'Version Data (JSON)'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, version in enumerate(queryset, 2):
            version_data = version.version_data
            booking = version.booking
            ws.cell(row=row_num, column=1, value=version.id)
            ws.cell(row=row_num, column=2, value=booking.id if booking else '')
            ws.cell(row=row_num, column=3, value=version_data.get('name', ''))
            ws.cell(row=row_num, column=4, value=version_data.get('room_no', ''))
            ws.cell(row=row_num, column=5, value=version.edited_by.username if version.edited_by else '')
            ws.cell(row=row_num, column=6, value=version.edited_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=7, value='Yes' if version.is_manager_edit else 'No')
            ws.cell(row=row_num, column=8, value=json.dumps(version_data, indent=2))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"booking_versions_export_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected booking versions to Excel"


@admin.register(RoomIssue)
class RoomIssueAdmin(admin.ModelAdmin):
    list_display = ['room', 'title', 'issue_type', 'status', 'priority', 'reported_by', 'reported_at', 'fixed_by', 'fixed_at']
    list_filter = ['status', 'issue_type', 'priority', 'reported_at']
    search_fields = ['room__room_number', 'title', 'description']
    readonly_fields = ['reported_at', 'created_at', 'updated_at']
    change_list_template = 'admin/bookings/roomissue/change_list.html'
    
    fieldsets = (
        ('Issue Information', {
            'fields': ('room', 'issue_type', 'title', 'description', 'priority', 'status')
        }),
        ('Reporting', {
            'fields': ('reported_by', 'reported_at')
        }),
        ('Resolution', {
            'fields': ('fixed_by', 'fixed_at', 'resolution_notes')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def save_model(self, request, obj, form, change):
        # Set reported_by if creating new issue
        if not change and not obj.reported_by:
            obj.reported_by = request.user
        
        # If status is being set to 'fixed' or 'resolved' and fixed_by/fixed_at not set
        if obj.status in ['fixed', 'resolved'] and not obj.fixed_at:
            obj.fixed_by = request.user
            obj.fixed_at = timezone.now()
        
        super().save_model(request, obj, form, change)
    
    actions = ['mark_as_fixed', 'mark_as_in_progress', 'export_excel']
    
    def mark_as_fixed(self, request, queryset):
        """Mark selected issues as fixed"""
        updated = queryset.update(
            status='fixed',
            fixed_by=request.user,
            fixed_at=timezone.now()
        )
        self.message_user(request, f'{updated} issue(s) marked as fixed.')
    mark_as_fixed.short_description = "Mark selected issues as fixed"
    
    def mark_as_in_progress(self, request, queryset):
        """Mark selected issues as in progress"""
        updated = queryset.update(status='in_progress')
        self.message_user(request, f'{updated} issue(s) marked as in progress.')
    mark_as_in_progress.short_description = "Mark selected issues as in progress"
    
    def export_excel(self, request, queryset):
        """Export room issues to Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Room Issues"
        
        # Headers
        headers = [
            'ID', 'Room Number', 'Room Type', 'Issue Type', 'Title',
            'Description', 'Status', 'Priority', 'Reported By',
            'Reported At', 'Fixed By', 'Fixed At', 'Resolution Notes',
            'Created At', 'Updated At'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, issue in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=issue.id)
            ws.cell(row=row_num, column=2, value=issue.room.room_number)
            ws.cell(row=row_num, column=3, value=issue.room.get_room_type_display())
            ws.cell(row=row_num, column=4, value=issue.get_issue_type_display())
            ws.cell(row=row_num, column=5, value=issue.title)
            ws.cell(row=row_num, column=6, value=issue.description)
            ws.cell(row=row_num, column=7, value=issue.get_status_display())
            ws.cell(row=row_num, column=8, value=issue.get_priority_display())
            ws.cell(row=row_num, column=9, value=issue.reported_by.username if issue.reported_by else '')
            ws.cell(row=row_num, column=10, value=issue.reported_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=11, value=issue.fixed_by.username if issue.fixed_by else '')
            ws.cell(row=row_num, column=12, value=issue.fixed_at.strftime('%Y-%m-%d %H:%M:%S') if issue.fixed_at else '')
            ws.cell(row=row_num, column=13, value=issue.resolution_notes)
            ws.cell(row=row_num, column=14, value=issue.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=15, value=issue.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"room_issues_export_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    export_excel.short_description = "Export selected room issues to Excel"

